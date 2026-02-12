"""Excel Generator Skill.

Built-in skill for professional Excel spreadsheet creation based on Pythinker SKILL.md.

Features from Pythinker four-layer implementation:
1. Structure Layer - Logical organization, easy navigation
2. Information Layer - Accurate, complete, insightful content
3. Visual Layer - Professional appearance, meaningful visuals
4. Interaction Layer - Usable immediately, clear actions

Key capabilities:
- Theme system (12 professional themes)
- Data blocks with proper borders
- Chart integration (bar, line, pie, sparklines)
- Conditional formatting
- Key insights sections
- Professional typography (serif + sans-serif pairing)
- Semantic colors for data meaning
- Sheet navigation with hyperlinks
"""

from dataclasses import dataclass, field
from enum import Enum
from typing import Any, ClassVar

from app.domain.models.skill import Skill, SkillCategory, SkillInvocationType, SkillSource


class ExcelTheme(str, Enum):
    """Available Excel themes."""

    ELEGANT_BLACK = "elegant_black"
    CORPORATE_BLUE = "corporate_blue"
    FOREST_GREEN = "forest_green"
    BURGUNDY = "burgundy"
    SLATE_GRAY = "slate_gray"
    NAVY = "navy"
    CHARCOAL = "charcoal"
    DEEP_PURPLE = "deep_purple"
    TEAL = "teal"
    WARM_BROWN = "warm_brown"
    ROYAL_BLUE = "royal_blue"
    OLIVE = "olive"


@dataclass
class ThemeConfig:
    """Theme configuration for Excel styling."""

    primary: str  # Primary color for headers
    light: str  # Light background color
    accent: str  # Accent color for highlights
    chart_colors: list[str] = field(default_factory=list)

    @property
    def header_fill(self) -> str:
        """Get header fill color."""
        return self.primary

    @property
    def alternate_row_fill(self) -> str:
        """Get alternate row fill color."""
        return self.light


# Theme configurations from Pythinker SKILL.md
THEMES: dict[ExcelTheme, ThemeConfig] = {
    ExcelTheme.ELEGANT_BLACK: ThemeConfig(
        primary="2D2D2D",
        light="E5E5E5",
        accent="2D2D2D",
        chart_colors=["2D2D2D", "4A4A4A", "6B6B6B", "8C8C8C", "ADADAD", "CFCFCF"],
    ),
    ExcelTheme.CORPORATE_BLUE: ThemeConfig(
        primary="1F4E79",
        light="D6E3F0",
        accent="1F4E79",
        chart_colors=["1F4E79", "2E75B6", "5B9BD5", "9DC3E6", "BDD7EE", "DEEBF7"],
    ),
    ExcelTheme.FOREST_GREEN: ThemeConfig(
        primary="2E5A4C",
        light="D4E5DE",
        accent="2E5A4C",
        chart_colors=["2E5A4C", "4A7C6B", "6B9D8A", "8CBEA9", "ADDFC8", "CEFFF7"],
    ),
    ExcelTheme.BURGUNDY: ThemeConfig(
        primary="722F37",
        light="E8D5D7",
        accent="722F37",
        chart_colors=["722F37", "944855", "B66173", "D87A91", "FA93AF", "FFACCD"],
    ),
    ExcelTheme.SLATE_GRAY: ThemeConfig(
        primary="4A5568",
        light="E2E8F0",
        accent="4A5568",
        chart_colors=["4A5568", "6B7280", "9CA3AF", "D1D5DB", "E5E7EB", "F3F4F6"],
    ),
    ExcelTheme.NAVY: ThemeConfig(
        primary="1E3A5F",
        light="D3DCE6",
        accent="1E3A5F",
        chart_colors=["1E3A5F", "2F5A8A", "407AB5", "519AE0", "62BAFF", "93CFFF"],
    ),
    ExcelTheme.CHARCOAL: ThemeConfig(
        primary="36454F",
        light="DDE1E4",
        accent="36454F",
        chart_colors=["36454F", "546E7A", "78909C", "90A4AE", "B0BEC5", "CFD8DC"],
    ),
    ExcelTheme.DEEP_PURPLE: ThemeConfig(
        primary="4A235A",
        light="E1D5E7",
        accent="4A235A",
        chart_colors=["4A235A", "6A3B7A", "8A539A", "AA6BBA", "CA83DA", "EA9BFA"],
    ),
    ExcelTheme.TEAL: ThemeConfig(
        primary="1A5F5F",
        light="D3E5E5",
        accent="1A5F5F",
        chart_colors=["1A5F5F", "2A7F7F", "3A9F9F", "4ABFBF", "5ADFDF", "6AFFFF"],
    ),
    ExcelTheme.WARM_BROWN: ThemeConfig(
        primary="5D4037",
        light="E6DDD9",
        accent="5D4037",
        chart_colors=["5D4037", "795548", "8D6E63", "A1887F", "BCAAA4", "D7CCC8"],
    ),
    ExcelTheme.ROYAL_BLUE: ThemeConfig(
        primary="1A237E",
        light="D3D5E8",
        accent="1A237E",
        chart_colors=["1A237E", "283593", "3949AB", "5C6BC0", "7986CB", "9FA8DA"],
    ),
    ExcelTheme.OLIVE: ThemeConfig(
        primary="556B2F",
        light="E0E5D5",
        accent="556B2F",
        chart_colors=["556B2F", "6B8E23", "9ACD32", "ADFF2F", "7CFC00", "7FFF00"],
    ),
}


# Semantic colors (independent of theme)
class SemanticColors:
    """Semantic colors for data meaning."""

    POSITIVE = "2E7D32"  # Green - growth, profit, success
    NEGATIVE = "C62828"  # Red - decline, loss, failure
    WARNING = "F57C00"  # Orange - caution, attention
    NEUTRAL = "757575"  # Gray - neutral data


# Row highlight colors for emphasis
class HighlightColors:
    """Row highlight colors for subtle distinction."""

    EMPHASIS = "E6F3FF"  # Light blue - top rated, important
    SECTION = "FFF3E0"  # Light orange - section dividers
    INPUT = "FFFDE7"  # Light yellow - editable cells
    SPECIAL = "FFF9C4"  # Yellow - base case, benchmarks
    SUCCESS = "E8F5E9"  # Light green - passed, completed
    WARNING = "FFCCBC"  # Light coral - needs attention


@dataclass
class FeatureMapping:
    """Feature to User Value mapping."""

    feature: str
    user_value: str
    when_to_use: str


@dataclass
class ExcelConfig:
    """Configuration for Excel generation."""

    theme: ExcelTheme = ExcelTheme.CORPORATE_BLUE
    include_overview: bool = True
    include_insights: bool = True
    include_charts: bool = True
    freeze_panes: bool = True
    add_filters: bool = True
    serif_font: str = "Source Serif Pro"
    sans_font: str = "Source Sans Pro"


class ExcelGeneratorSkill:
    """Excel generator skill implementation.

    Goal: Make the user able to use the Excel immediately and gain insights upon opening.

    Core Principle: Enrich visuals as much as possible, while ensuring content clarity
    and not adding cognitive burden. Every visual element should be meaningful and
    purposeful—serving the content, not decorating it.
    """

    # Feature categories from Pythinker SKILL.md
    FEATURE_CATEGORIES: ClassVar[dict[str, list[FeatureMapping]]] = {
        "understand_data": [
            FeatureMapping("Bar/Column Chart", "See comparisons at a glance", "Comparing values across categories"),
            FeatureMapping("Line Chart", "See trends at a glance", "Time series data"),
            FeatureMapping("Pie Chart", "See proportions at a glance", "Part-to-whole (≤6 categories)"),
            FeatureMapping("Data Bars", "Compare magnitude without leaving cell", "Numeric columns needing comparison"),
            FeatureMapping("Color Scale", "Heatmap effect, patterns pop out", "Matrices, ranges, distributions"),
            FeatureMapping("Sparklines", "See trend within a single cell", "Summary rows with historical context"),
        ],
        "find_what_matters": [
            FeatureMapping("Pre-sorting", "Most important data comes first", "Rankings, Top N, priorities"),
            FeatureMapping(
                "Conditional Highlighting", "Key data stands out automatically", "Outliers, thresholds, Top/Bottom N"
            ),
            FeatureMapping("Icon Sets", "Status visible at a glance", "KPI status, categorical states"),
            FeatureMapping("Bold/Color Emphasis", "Visual distinction", "Summary rows, key metrics"),
            FeatureMapping("KEY INSIGHTS Section", "Conclusions delivered directly", "Analytical reports"),
        ],
        "save_time": [
            FeatureMapping("Overview Sheet", "Summary on first page", "All multi-sheet files"),
            FeatureMapping("Pre-calculated Summaries", "Results ready", "Data requiring statistics"),
            FeatureMapping("Consistent Number Formats", "No format adjustments needed", "All numeric data"),
            FeatureMapping("Freeze Panes", "Headers visible while scrolling", "Tables with >10 rows"),
            FeatureMapping("Sheet Index with Links", "Quick navigation", "Files with >3 sheets"),
        ],
        "use_directly": [
            FeatureMapping("Filters", "Users can explore data", "Exploratory analysis"),
            FeatureMapping("Hyperlinks", "Click to navigate", "Cross-sheet references"),
            FeatureMapping("Print-friendly Layout", "Ready to print/PDF", "Reports for sharing"),
            FeatureMapping("Formulas (not hardcoded)", "Results update", "Models, forecasts"),
            FeatureMapping("Data Validation Dropdowns", "Prevent input errors", "Templates for user input"),
        ],
        "trust_data": [
            FeatureMapping("Data Source Attribution", "Know where data comes from", "All external data"),
            FeatureMapping("Generation Date", "Know data freshness", "Time-sensitive reports"),
            FeatureMapping("Data Time Range", "Know what period covered", "Time series data"),
            FeatureMapping("Professional Formatting", "Looks reliable", "External-facing files"),
            FeatureMapping("Consistent Precision", "No doubts about accuracy", "All numeric values"),
        ],
        "gain_insights": [
            FeatureMapping("Comparison Columns (Δ, %)", "No manual calculation", "YoY, MoM, A vs B"),
            FeatureMapping("Rank Column", "Position visible directly", "Competitive analysis"),
            FeatureMapping("Grouped Summaries", "Aggregated by dimension", "Segmented analysis"),
            FeatureMapping("Trend Indicators (↑↓)", "Direction clear at a glance", "Change direction matters"),
            FeatureMapping("Insight Text", "The 'so what' is stated", "Analytical reports"),
        ],
    }

    @classmethod
    def get_skill_definition(cls) -> Skill:
        """Get the skill definition for registration."""
        return Skill(
            id="excel-generator",
            name="Excel Generator",
            description=(
                "Professional Excel spreadsheet creation with a focus on aesthetics and data analysis. "
                "Use when creating spreadsheets for organizing, analyzing, and presenting structured data "
                "in a clear and professional format."
            ),
            category=SkillCategory.DATA_ANALYSIS,
            source=SkillSource.OFFICIAL,
            icon="table-2",
            required_tools=["code_execute_python", "file_write"],
            optional_tools=["file_read"],
            system_prompt_addition=cls._get_system_prompt(),
            invocation_type=SkillInvocationType.BOTH,
            trigger_patterns=[
                r"(?i)create.*excel",
                r"(?i)generate.*spreadsheet",
                r"(?i)make.*xlsx",
                r"(?i)build.*excel",
            ],
            default_enabled=False,
            version="1.0.0",
            author="Pythinker",
        )

    @classmethod
    def _get_system_prompt(cls) -> str:
        """Get the system prompt addition for this skill."""
        return """# Excel Generator Skill

## Goal
Make the user able to use the Excel immediately and gain insights upon opening.

## Core Principle
Enrich visuals as much as possible, while ensuring content clarity and not adding cognitive burden.
Every visual element should be meaningful and purposeful—serving the content, not decorating it.

## User Needs & Feature Matching

Before creating any Excel, think through:
1. **What does the user need?** — Not "an Excel file", but what problem are they solving?
2. **What can I provide?** — Which features will help them?
3. **How to match?** — Select the right combination for this specific scenario.

### Feature Categories

**Help Users「Understand Data」**
- Bar/Column Chart: See comparisons at a glance (for comparing values across categories)
- Line Chart: See trends at a glance (for time series data)
- Pie Chart: See proportions at a glance (for part-to-whole, ≤6 categories)
- Data Bars: Compare magnitude without leaving the cell
- Color Scale: Heatmap effect, patterns pop out
- Sparklines: See trend within a single cell

**Help Users「Find What Matters」**
- Pre-sorting: Most important data comes first
- Conditional Highlighting: Key data stands out automatically
- Icon Sets: Status visible at a glance (use sparingly)
- Bold/Color Emphasis: Visual distinction between primary and secondary
- KEY INSIGHTS Section: Conclusions delivered directly

**Help Users「Save Time」**
- Overview Sheet: Summary on first page (for all multi-sheet files)
- Pre-calculated Summaries: Results ready
- Consistent Number Formats: No format adjustments needed
- Freeze Panes: Headers visible while scrolling (for tables with >10 rows)
- Sheet Index with Links: Quick navigation (for files with >3 sheets)

## Four-Layer Implementation

### Layer 1: Structure
- Sheets: 3-5 ideal, max 7
- First sheet: Always "Overview" with summary and navigation
- Sheet order: General → Specific (Overview → Data → Analysis)
- Left margin: Column A empty (width 3)
- Content start: Cell B2
- Section spacing: 1 empty row between sections

### Layer 2: Information
- Every numeric cell must have `number_format` set
- Same column = same precision
- Include data source, time range, generation date
- Add KEY INSIGHTS section for analytical content

### Layer 3: Visual
- Hide gridlines: `ws.sheet_view.showGridLines = False`
- Use theme colors consistently
- Typography: Serif for titles (Source Serif Pro), Sans-serif for data (Source Sans Pro)
- Data blocks get outer frame borders

### Layer 4: Interaction
- Add filters for exploratory analysis
- Include hyperlinks for navigation
- Freeze panes for headers
- Use formulas instead of hardcoded values

## Implementation Notes

Use `openpyxl` library for Excel generation. Key patterns:

```python
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart

# Theme colors
THEME = {
    'primary': '1F4E79',
    'light': 'D6E3F0',
    'accent': '1F4E79',
}

# Hide gridlines
ws.sheet_view.showGridLines = False

# Header style
header_fill = PatternFill(start_color=THEME['primary'], fill_type='solid')
header_font = Font(name='Source Serif Pro', bold=True, color='FFFFFF')

# Number formats
'#,##0'      # Integer: 1,234,567
'#,##0.00'   # Decimal: 1,234.56
'0.0%'       # Percentage: 12.3%
'$#,##0.00'  # Currency: $1,234.56
```

Always ensure the output is immediately usable and professionally formatted.
"""

    @classmethod
    def generate_openpyxl_code(
        cls,
        data: list[dict[str, Any]],
        title: str,
        config: ExcelConfig | None = None,
    ) -> str:
        """Generate openpyxl Python code for Excel creation.

        This generates the code that the agent should execute to create the Excel file.
        """
        if config is None:
            config = ExcelConfig()

        theme = THEMES[config.theme]

        # Build the code string
        return f'''
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from datetime import datetime

# Theme configuration
THEME = {{
    'primary': '{theme.primary}',
    'light': '{theme.light}',
    'accent': '{theme.accent}',
}}

SERIF_FONT = '{config.serif_font}'
SANS_FONT = '{config.sans_font}'

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Data"

# Hide gridlines
ws.sheet_view.showGridLines = False

# Set column A as margin
ws.column_dimensions['A'].width = 3

# Styles
header_fill = PatternFill(start_color=THEME['primary'], fill_type='solid')
header_font = Font(name=SERIF_FONT, size=11, bold=True, color='FFFFFF')
title_font = Font(name=SERIF_FONT, size=18, bold=True, color=THEME['primary'])
data_font = Font(name=SANS_FONT, size=11)

# Title
ws['B2'] = "{title}"
ws['B2'].font = title_font

# Headers and data would be added here based on the input data
# This is a template - actual implementation would iterate through data

# Save workbook
wb.save("output.xlsx")
print("Excel file created: output.xlsx")
'''

    @classmethod
    def get_number_format(cls, data_type: str) -> str:
        """Get the appropriate number format for a data type."""
        formats = {
            "integer": "#,##0",
            "decimal": "#,##0.00",
            "decimal1": "#,##0.0",
            "percentage": "0.0%",
            "currency": "$#,##0.00",
            "currency_eur": "€#,##0.00",
            "date": "YYYY-MM-DD",
            "datetime": "YYYY-MM-DD HH:MM",
        }
        return formats.get(data_type, "General")

    @classmethod
    def get_semantic_color(cls, value_type: str) -> str:
        """Get semantic color for data meaning."""
        colors = {
            "positive": SemanticColors.POSITIVE,
            "negative": SemanticColors.NEGATIVE,
            "warning": SemanticColors.WARNING,
            "neutral": SemanticColors.NEUTRAL,
        }
        return colors.get(value_type, SemanticColors.NEUTRAL)

    @classmethod
    def get_highlight_color(cls, highlight_type: str) -> str:
        """Get highlight color for row emphasis."""
        colors = {
            "emphasis": HighlightColors.EMPHASIS,
            "section": HighlightColors.SECTION,
            "input": HighlightColors.INPUT,
            "special": HighlightColors.SPECIAL,
            "success": HighlightColors.SUCCESS,
            "warning": HighlightColors.WARNING,
        }
        return colors.get(highlight_type, HighlightColors.EMPHASIS)
