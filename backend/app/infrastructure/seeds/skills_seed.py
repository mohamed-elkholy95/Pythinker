"""Official skills seed data.

This module defines the built-in skills that come with Pythinker.
Skills are prepackaged capabilities that group related tools together.

Implements Claude's skills architecture patterns:
- Invocation control (user, ai, both)
- Tool restrictions (allowed_tools)
- Dynamic context injection support
- Trigger patterns for automatic activation
"""

import contextlib
from datetime import UTC, datetime
from pathlib import Path

from app.domain.models.skill import InstructionTrustLevel, Skill, SkillCategory, SkillInvocationType, SkillSource

# =============================================================================
# SKILL CREATOR PATHS
# =============================================================================

# Path to the skill-creator resources (relative to this file)
SKILL_CREATOR_DIR = Path(__file__).parent / "skills" / "skill-creator"
SKILL_CREATOR_SKILL_MD = SKILL_CREATOR_DIR / "SKILL.md"

# Workspace path where skill-creator will be available to agents
WORKSPACE_SKILL_CREATOR_PATH = "/workspace/skills/skill-creator"


def get_skill_creator_content() -> str:
    """Get the content of the skill-creator SKILL.md file.

    Returns:
        Content of SKILL.md or empty string if not found
    """
    if SKILL_CREATOR_SKILL_MD.exists():
        return SKILL_CREATOR_SKILL_MD.read_text()
    return ""


def get_skill_creator_resources() -> dict[str, str]:
    """Get all skill-creator resource files.

    Returns:
        Dict mapping relative paths to file contents
    """
    resources = {}
    if not SKILL_CREATOR_DIR.exists():
        return resources

    for file_path in SKILL_CREATOR_DIR.rglob("*"):
        if file_path.is_file():
            rel_path = file_path.relative_to(SKILL_CREATOR_DIR)
            with contextlib.suppress(UnicodeDecodeError):
                resources[str(rel_path)] = file_path.read_text()
    return resources


# =============================================================================
# OFFICIAL SKILLS DEFINITIONS
# =============================================================================

OFFICIAL_SKILLS: list[Skill] = [
    # -------------------------------------------------------------------------
    # Research Skill
    # -------------------------------------------------------------------------
    Skill(
        id="research",
        name="Research",
        description="Web research and information gathering. Search the web, browse sites, and compile findings.",
        category=SkillCategory.RESEARCH,
        source=SkillSource.OFFICIAL,
        icon="search",
        required_tools=[
            "info_search_web",
            "browser_navigate",
            "browser_get_content",
            "file_write",
        ],
        optional_tools=[
            "browser_agent_run",
            "browser_agent_extract",
        ],
        system_prompt_addition="""<research_skill>
## MANDATORY Research Workflow

### Step 1: Search (info_search_web)
- Find relevant sources with search queries
- Note the URLs returned in search results

### Step 2: BROWSE ACTUAL PAGES (CRITICAL)
- You MUST use browser_navigate or browser_get_content to visit URLs
- Do NOT rely on search snippets - they are often outdated or incomplete
- Extract actual content from the pages you visit
- Visit at least 3-5 sources for comprehensive research

### Step 3: Extract & Verify
- Pull specific data, specs, and facts from page content
- Cross-reference claims across multiple sources
- Note the current date context for time-sensitive information

### Step 4: Synthesize with Citations
- Compile findings from ACTUAL page content (not LLM knowledge)
- Include inline citations [1], [2] linking to visited URLs
- End with a References section listing all sources

## CRITICAL RULES
- NEVER generate research reports from search snippets alone
- ALWAYS browse to actual URLs before writing conclusions
- If browser_navigate fails, use browser_get_content as fallback
- Cite only pages you actually visited and extracted content from
</research_skill>""",
        default_enabled=True,
        version="1.0.0",
        author="Pythinker",
        is_premium=False,
        # Claude-style configuration
        invocation_type=SkillInvocationType.BOTH,
        trigger_patterns=[r"research\s+", r"find\s+information", r"look\s+up"],
        instruction_trust_level=InstructionTrustLevel.SYSTEM_AUTHORED,
    ),
    # -------------------------------------------------------------------------
    # Coding Skill
    # -------------------------------------------------------------------------
    Skill(
        id="coding",
        name="Coding",
        description="Code generation, debugging, and refactoring. Read, write, and execute code.",
        category=SkillCategory.CODING,
        source=SkillSource.OFFICIAL,
        icon="code",
        required_tools=[
            "file_read",
            "file_write",
            "file_str_replace",
            "file_find_in_content",
            "file_find_by_name",
            "shell_exec",
            "code_execute",
        ],
        optional_tools=[
            "code_execute_python",
            "code_execute_javascript",
        ],
        system_prompt_addition="""<coding_skill>
When writing or modifying code:

1. **Understand First**: Always read existing code before modifying
   - Use file_read to examine files
   - Use file_find_in_content to search for patterns
   - Understand the codebase structure

2. **Make Focused Changes**: Don't over-engineer
   - Only change what's necessary
   - Don't add features unless asked
   - Keep solutions simple

3. **Code Quality**:
   - Follow language-specific best practices
   - Handle errors gracefully
   - Write self-documenting code
   - Add comments only where logic isn't obvious

4. **Verify Changes**:
   - Test code before considering task complete
   - Use code_execute to run and verify
   - Check for syntax errors and edge cases

5. **Version Control**:
   - Stage specific files, not everything
   - Write clear commit messages
   - Don't commit sensitive data
</coding_skill>""",
        default_enabled=True,
        version="1.0.0",
        author="Pythinker",
        is_premium=False,
        # Claude-style configuration
        invocation_type=SkillInvocationType.BOTH,
        trigger_patterns=[r"write\s+code", r"debug", r"fix\s+", r"implement", r"refactor"],
        instruction_trust_level=InstructionTrustLevel.SYSTEM_AUTHORED,
    ),
    # -------------------------------------------------------------------------
    # Browser Skill
    # -------------------------------------------------------------------------
    Skill(
        id="browser",
        name="Browser",
        description="Web browsing and interaction. Navigate websites, fill forms, click buttons.",
        category=SkillCategory.BROWSER,
        source=SkillSource.OFFICIAL,
        icon="globe",
        required_tools=[
            "browser_navigate",
            "browser_view",
            "browser_click",
            "browser_input",
            "browser_scroll_down",
            "browser_scroll_up",
            "browser_get_content",
        ],
        optional_tools=[
            "browser_restart",
            "browser_press_key",
            "browser_select_option",
            "browser_move_mouse",
            "browser_console_exec",
            "browser_console_view",
        ],
        default_enabled=True,
        version="1.0.0",
        author="Pythinker",
        is_premium=False,
        instruction_trust_level=InstructionTrustLevel.SYSTEM_AUTHORED,
    ),
    # -------------------------------------------------------------------------
    # File Management Skill
    # -------------------------------------------------------------------------
    Skill(
        id="file-management",
        name="File Management",
        description="File operations and organization. Create, read, edit, and organize files.",
        category=SkillCategory.FILE_MANAGEMENT,
        source=SkillSource.OFFICIAL,
        icon="folder",
        required_tools=[
            "file_read",
            "file_write",
            "file_str_replace",
            "file_find_in_content",
            "file_find_by_name",
        ],
        optional_tools=[
            "shell_exec",
        ],
        default_enabled=True,
        version="1.0.0",
        author="Pythinker",
        is_premium=False,
        instruction_trust_level=InstructionTrustLevel.SYSTEM_AUTHORED,
    ),
    # -------------------------------------------------------------------------
    # Data Analysis Skill
    # -------------------------------------------------------------------------
    Skill(
        id="data-analysis",
        name="Data Analysis",
        description="Data processing, analysis, and visualization. Work with CSV, JSON, and create charts. Use for: analyzing datasets, creating visualizations, statistical analysis.",
        category=SkillCategory.DATA_ANALYSIS,
        source=SkillSource.OFFICIAL,
        icon="bar-chart",
        required_tools=[
            "file_read",
            "file_write",
            "code_execute_python",
            "shell_exec",
        ],
        optional_tools=[
            "code_execute",
        ],
        system_prompt_addition="""<data_analysis>
# Data Analysis Skill

## CRITICAL: Execute Code, Don't Just Describe
- You MUST use code_execute_python to perform actual analysis
- DO NOT write explanatory text without running code first
- ALWAYS save outputs (charts, processed data) to /workspace/

## Workflow
1. **Load Data**: Read CSV/JSON/Excel files using pandas
2. **Explore**: Run df.info(), df.describe(), df.head()
3. **Clean**: Handle missing values, fix data types
4. **Analyze**: Calculate statistics, find patterns
5. **Visualize**: Create charts with matplotlib/seaborn
6. **Export**: Save results to files

## Required Output
- Summary statistics (mean, median, std, etc.)
- Visualizations saved as PNG/SVG files
- Processed data saved to CSV/Excel if requested

## Python Pattern
```python
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns

# Load data
df = pd.read_csv('/workspace/data.csv')

# Analysis
print(df.describe())

# Visualization
plt.figure(figsize=(10, 6))
sns.barplot(data=df, x='category', y='value')
plt.title('Analysis Result')
plt.savefig('/workspace/chart.png', dpi=150, bbox_inches='tight')
plt.close()
print("Chart saved to /workspace/chart.png")
```
</data_analysis>""",
        default_enabled=False,
        version="1.0.0",
        author="Pythinker",
        is_premium=False,
        invocation_type=SkillInvocationType.BOTH,
        trigger_patterns=[
            r"analyze\s+data",
            r"data\s+analysis",
            r"create\s+chart",
            r"visualize",
            r"statistics",
        ],
        instruction_trust_level=InstructionTrustLevel.SYSTEM_AUTHORED,
    ),
    # -------------------------------------------------------------------------
    # Excel Generator Skill
    # -------------------------------------------------------------------------
    Skill(
        id="excel-generator",
        name="Excel Generator",
        description="Create professional Excel spreadsheets with formatting, formulas, and charts. Use for: generating reports as Excel, data exports, spreadsheet creation.",
        category=SkillCategory.DATA_ANALYSIS,
        source=SkillSource.OFFICIAL,
        icon="file-spreadsheet",
        required_tools=[
            "code_execute_python",
            "file_write",
        ],
        optional_tools=[
            "file_read",
            "shell_exec",
        ],
        # Restrict to Python execution to ensure Excel file is generated
        allowed_tools=[
            "code_execute_python",
            "file_write",
            "file_read",
            "shell_exec",
            "message_notify_user",
            "message_ask_user",
            "idle_standby",
        ],
        system_prompt_addition="""<excel_generator>
# Excel Generator Skill - MANDATORY RULES

## CRITICAL: Output Format Requirement
**YOU MUST create an actual .xlsx Excel file using Python code.**
- DO NOT write markdown reports or plain text summaries
- DO NOT provide data as formatted text
- ALWAYS execute Python code to generate a real Excel file

## Workflow
1. **Understand Requirements**: Clarify what data/columns the user needs
2. **Write Python Script**: Use openpyxl or xlsxwriter
3. **Execute Code**: Use code_execute_python to create the .xlsx file
4. **Verify File**: Confirm the Excel file was created successfully
5. **Notify User**: Tell user the file location and what's in it

## Required Python Code Pattern
```python
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill

wb = Workbook()
ws = wb.active
ws.title = "Sheet1"

# Add headers with formatting
headers = ["Column1", "Column2", "Column3"]
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill

# Add data rows
data = [
    ["Value1", "Value2", "Value3"],
]
for row_idx, row_data in enumerate(data, 2):
    for col_idx, value in enumerate(row_data, 1):
        ws.cell(row=row_idx, column=col_idx, value=value)

# Save file
wb.save("/workspace/output.xlsx")
print("Excel file created: /workspace/output.xlsx")
```

## Formatting Requirements
- Headers: Bold, colored background, centered
- Borders: Thin borders around all cells
- Column widths: Auto-fit or reasonable defaults
- Number formats: Apply appropriate formats (currency, percentage, dates)

## DO NOT
- Write text-only responses explaining data
- Output markdown tables instead of Excel files
- Skip the code execution step
- Forget to save the file to /workspace/
</excel_generator>""",
        default_enabled=False,
        version="1.0.0",
        author="Pythinker",
        is_premium=False,
        invocation_type=SkillInvocationType.BOTH,
        trigger_patterns=[
            r"excel\s+",
            r"spreadsheet",
            r"\.xlsx",
            r"create\s+.*\s+excel",
            r"generate\s+.*\s+excel",
            r"export\s+to\s+excel",
        ],
        instruction_trust_level=InstructionTrustLevel.SYSTEM_AUTHORED,
    ),
    # -------------------------------------------------------------------------
    # Stock Analysis Skill (Premium)
    # -------------------------------------------------------------------------
    Skill(
        id="stock-analysis",
        name="Stock Analysis",
        description="Analyze stocks, financial data, and market trends. Get quotes, charts, and insights.",
        category=SkillCategory.DATA_ANALYSIS,
        source=SkillSource.OFFICIAL,
        icon="trending-up",
        required_tools=[
            "info_search_web",
            "browser_navigate",
            "code_execute_python",
            "file_write",
        ],
        optional_tools=[
            "browser_agent_run",
            "browser_agent_extract",
        ],
        system_prompt_addition="""When analyzing stocks:
- Use yfinance for historical data
- Calculate technical indicators (MA, RSI, MACD)
- Create price charts with matplotlib
- Provide clear buy/sell/hold recommendations with reasoning
- Include risk warnings and disclaimers""",
        default_enabled=False,
        version="1.0.0",
        author="Pythinker",
        is_premium=True,
        instruction_trust_level=InstructionTrustLevel.SYSTEM_AUTHORED,
    ),
    # -------------------------------------------------------------------------
    # Web Pilot Skill (Autonomous browsing)
    # -------------------------------------------------------------------------
    Skill(
        id="web-pilot",
        name="Web Pilot",
        description="Autonomous web browsing with AI. Navigate complex sites, fill forms, and extract data.",
        category=SkillCategory.BROWSER,
        source=SkillSource.OFFICIAL,
        icon="bot",
        required_tools=[
            "browser_agent_run",
            "browser_agent_extract",
        ],
        optional_tools=[
            "browser_navigate",
            "browser_view",
            "file_write",
        ],
        system_prompt_addition="""When using Web Pilot:
- Describe the browsing goal clearly
- Use extract for structured data retrieval
- Handle multi-step workflows automatically""",
        default_enabled=False,
        version="1.0.0",
        author="Pythinker",
        is_premium=False,
        instruction_trust_level=InstructionTrustLevel.SYSTEM_AUTHORED,
    ),
    # -------------------------------------------------------------------------
    # Skill Creator Meta-Skill (for creating custom skills)
    # -------------------------------------------------------------------------
    Skill(
        id="skill-creator",
        name="Skill Creator",
        description="Guide for creating effective skills. Use when users want to create a new skill (or update an existing skill) that extends Pythinker's capabilities with specialized knowledge, workflows, or tool integrations.",
        category=SkillCategory.CUSTOM,
        source=SkillSource.OFFICIAL,
        icon="puzzle",
        required_tools=[
            "message_ask_user",
            "message_notify_user",
            "skill_create",
            "file_write",
            "file_read",
            "shell_exec",
        ],
        optional_tools=[
            "skill_list_user",
            "skill_delete",
            "file_find_by_name",
            "code_execute_python",
        ],
        invocation_type=SkillInvocationType.BOTH,
        trigger_patterns=[
            r"create\s+(?:a\s+)?skill",
            r"make\s+(?:a\s+)?skill",
            r"build\s+(?:a\s+)?skill",
            r"new\s+skill",
        ],
        system_prompt_addition="""<skill_creator>
# Skill Creator

## IMPORTANT: First Step - Load Skill Creator Guide

When a user asks to create a skill, you MUST:

1. **First**, respond with: "I'll help you create the **[skill name]** skill. Let me first review the skill creation guidelines."

2. **Then**, use `file_read` to load and display the skill creator guide:
   ```
   file_read(path="/workspace/skills/skill-creator/SKILL.md")
   ```

   If the file doesn't exist at that path, the guide content is embedded below.

3. **Show the user** that you are reviewing the guidelines (this builds trust and transparency).

4. **Then** proceed with the skill creation process.

---

This skill provides guidance for creating effective skills.

## About Skills

Skills are modular, self-contained packages that extend Pythinker's capabilities by providing specialized knowledge, workflows, and tools. Think of them as "onboarding guides" for specific domains or tasks—they transform the agent from a general-purpose assistant into a specialized agent equipped with procedural knowledge that no model can fully possess.

### What Skills Provide

1. **Specialized workflows** - Multi-step procedures for specific domains
2. **Tool integrations** - Instructions for working with specific file formats or APIs
3. **Domain expertise** - Company-specific knowledge, schemas, business logic
4. **Bundled resources** - Scripts, references, and assets for complex and repetitive tasks

## Core Principles

### Concise is Key

The context window is a public good. Skills share the context window with everything else the agent needs: system prompt, conversation history, other skills' metadata, and the actual user request.

**Default assumption: The agent is already very smart.** Only add context the agent doesn't already have. Challenge each piece of information: "Does the agent really need this explanation?" and "Does this paragraph justify its token cost?"

Prefer concise examples over verbose explanations.

### Set Appropriate Degrees of Freedom

Match the level of specificity to the task's fragility and variability:

- **High freedom (text-based instructions)**: Multiple approaches valid, context-dependent decisions
- **Medium freedom (pseudocode/scripts with parameters)**: Preferred pattern exists, some variation OK
- **Low freedom (specific scripts, few parameters)**: Operations fragile, consistency critical

### Anatomy of a Skill

```
skill-name/
├── SKILL.md (required)
│   ├── YAML frontmatter (name, description)
│   └── Markdown instructions
└── Bundled Resources (optional)
    ├── scripts/      - Executable code (Python/Bash)
    ├── references/   - Documentation loaded as needed
    └── templates/    - Output assets (not loaded into context)
```

**Progressive Disclosure:**
1. **Metadata** - Always in context (~100 words)
2. **SKILL.md body** - When skill triggers (<500 lines)
3. **Bundled resources** - As needed

## Skill Creation Process

### Step 1: Understand the Skill with Concrete Examples

Skip only if usage patterns are already clearly understood.

Use `message_ask_user` to gather concrete examples:

```
message_ask_user(text="What functionality should this skill support? Can you give me examples of how you'd use it?")
```

**CRITICAL: Always use `message_ask_user` tool.** Without it, the system won't pause for user response.

### Step 2: Plan Reusable Skill Contents

Identify reusable resources for each use case:

| Resource Type | When to Use                     | Example                               |
|--------------|----------------------------------|---------------------------------------|
| `scripts/`   | Code rewritten repeatedly       | `rotate_pdf.py` for PDF rotation      |
| `templates/` | Same boilerplate each time      | HTML starter for webapp builder       |
| `references/`| Documentation needed repeatedly | Database schemas, API docs            |

Present the plan and use `message_ask_user` to confirm approach.

### Step 3: Define Skill Configuration

**Frontmatter:**
- `name`: Skill name (2-100 characters)
- `description`: MUST include what skill does AND when to use it
  - Good: "Code review with security focus. Use for: reviewing PRs, auditing security, analyzing code quality."
  - Bad: "Helps with code"

**Icon:** pen-tool, search, code, globe, folder, bar-chart, file-spreadsheet, trending-up, bot, sparkles, wand-2, file-text, puzzle, book-open, zap, shield

**Tools:** Select based on use case:
- **Search:** `info_search_web`
- **Browser:** `browser_navigate`, `browser_view`, `browser_get_content`, `browser_click`, `browser_input`, `browser_agent_run`, `browser_agent_extract`
- **Files:** `file_read`, `file_write`, `file_str_replace`, `file_find_in_content`, `file_find_by_name`
- **Code:** `code_execute`, `code_execute_python`, `code_execute_javascript`
- **Shell:** `shell_exec`
- **Communication:** `message_notify_user`, `message_ask_user`

### Step 4: Write SKILL.md Instructions

**Use imperative form:** "Read the file" not "You should read the file"

**Structure:**
```markdown
# [Skill Name]

## Overview
[One sentence: what this skill does]

## When to Use
[Specific triggers for this skill]

## Workflow
1. [First action]
2. [Second action]
3. [Third action]

## Guidelines
- [Specific guideline 1]
- [Specific guideline 2]

## Output Format
[How results should be presented]

## Example
Input: [example request]
Output: [example response]
```

**Keep under 4000 characters.** Focus on procedural "how to" rather than general knowledge.

### Step 5: Review and Confirm

Present complete skill definition and use `message_ask_user`:

```
message_ask_user(text="Does this skill definition look correct? Reply 'yes' to create it, or tell me what changes you'd like.")
```

**Wait for explicit user approval.**

### Step 6: Create and Deliver the Skill

**Only after user explicitly approves:**

Call `skill_create` with all parameters:

```python
skill_create(
  name="Human Blog Writer",
  description="Generates high-quality blog content with SEO optimization. Use for: writing blog posts, content outlines, SEO-optimized articles.",
  icon="pen-tool",
  required_tools=["file_write", "info_search_web", "browser_get_content"],
  system_prompt_addition="<human_blog_writer>\\n# Human Blog Writer\\n\\n## Overview\\n...\\n</human_blog_writer>",
  scripts=[
    {"filename": "seo_analyzer.py", "content": "# SEO analysis script\\nimport re\\n..."}
  ],
  references=[
    {"filename": "style_guide.md", "content": "# Writing Style Guide\\n..."},
    {"filename": "seo_best_practices.md", "content": "# SEO Best Practices\\n..."}
  ],
  templates=[
    {"filename": "blog_post_outline.md", "content": "# Blog Post Template\\n..."}
  ]
)
```

The system will automatically:
1. Validate skill structure
2. Package into `.skill` file
3. Display skill viewer with options: Add to My Skills, Download, Preview

## Design Patterns

For complex skills, consult these reference guides:
- **Sequential workflows**: `file_read(path="/workspace/skills/skill-creator/references/workflows.md")`
- **Output format patterns**: `file_read(path="/workspace/skills/skill-creator/references/output-patterns.md")`
- **Progressive disclosure**: `file_read(path="/workspace/skills/skill-creator/references/progressive-disclosure-patterns.md")`

### Sequential Workflow
```markdown
## Workflow
1. **Preparation**: [Setup actions]
2. **Execution**: [Main task actions]
3. **Verification**: [Quality checks]
4. **Delivery**: [Output actions]
```

### Conditional Logic
```markdown
## Decision Points
- If [condition A]: [action A]
- If [condition B]: [action B]
- Default: [fallback action]
```

### Progressive Disclosure (for complex skills)
Keep SKILL.md under 500 lines. When a skill supports multiple domains or variants:
- Put core workflow in SKILL.md
- Move variant-specific details to `references/` files
- Reference them clearly: "For AWS deployment, see references/aws.md"

### Output Template
```markdown
## Output Format
### Summary
[1-2 sentence overview]

### Findings
- [Finding 1]
- [Finding 2]

### Recommendations
[Actionable next steps]
```

## Quality Checklist

Before calling `skill_create`, verify:
- [ ] Description says WHAT the skill does AND WHEN to use it
- [ ] SKILL.md body uses imperative form ("Read the file" not "You should read")
- [ ] Skill is under 4000 characters (system_prompt_addition)
- [ ] No generic knowledge the agent already has — only procedural or domain-specific content
- [ ] Scripts tested via `code_execute_python` if included
- [ ] Example input/output pairs included where output quality matters

## CRITICAL RULES

1. **ALWAYS use `message_ask_user` tool** - Never write questions as plain text
2. **Wait for each response** before proceeding to next step
3. **Never create without approval** - Explicit user confirmation required
4. **Be concise** - Context window is a shared resource
5. **One question at a time** - Don't overwhelm with multiple questions
6. **Create bundled resources** - Scripts, references, templates for reusable content
7. **CALL skill_create** at the end - This packages and delivers the skill
8. **Load design pattern references** when the skill involves workflows, output formats, or multi-domain content

</skill_creator>""",
        default_enabled=False,
        version="1.0.0",
        author="Pythinker",
        is_premium=False,
        instruction_trust_level=InstructionTrustLevel.SYSTEM_AUTHORED,
    ),
    # -------------------------------------------------------------------------
    # Deal Finder Skill
    # -------------------------------------------------------------------------
    Skill(
        id="deal-finder",
        name="Deal Finder",
        description="Find the best deals, compare prices across retailers, discover coupons, and produce structured savings reports. Use when the user wants to buy a product, compare prices, find discounts, or get purchase recommendations.",
        category=SkillCategory.RESEARCH,
        source=SkillSource.OFFICIAL,
        icon="trending-up",
        required_tools=[
            "info_search_web",
            "browser_navigate",
            "browser_get_content",
            "file_write",
        ],
        optional_tools=[
            "browser_agent_run",
            "browser_agent_extract",
            "code_execute_python",
        ],
        system_prompt_addition="""<deal_finder_skill>
## Purpose

Find the best deal for a user's desired product by searching multiple sources, comparing prices, discovering coupons, and producing a structured recommendation with savings analysis.

## Workflow

1. **Parse the request.** Extract: product name, category, key specs, budget (if stated), preferred retailers, and region/country.
2. **Generate search queries.** Create 3-5 variant queries: exact product name, product + "best price", product + "deal OR coupon OR discount", product + "site:reddit.com OR site:slickdeals.net", and a generic category query for alternatives.
3. **Search multiple sources.** Execute all queries via web search. Prioritize results from: retailer sites (Amazon, Walmart, Best Buy, Newegg), price-comparison engines (Google Shopping, PriceGrabber, CamelCamelCamel), deal aggregators (Slickdeals, DealNews, Wirecutter), and community sources (Reddit r/deals, r/buildapcsales).
4. **Extract pricing data.** For each result, navigate to the product page using the browser. Collect: product title, current price, original/list price, seller, availability, shipping cost, rating, review count, and URL. Normalize all prices to the same currency.
5. **Find coupons and promo codes.** Search "[store name] promo code [current month/year]" for each retailer found. Check coupon aggregators: RetailMeNot, Honey, Coupons.com. Look for patterns: WELCOME10, SAVE20, FREESHIP, seasonal codes, newsletter signup discounts. Note eligibility constraints.
6. **Calculate total cost.** For each option compute: final_price = sale_price - coupon_discount + shipping + tax_estimate. Compute savings_vs_msrp and savings_percentage.
7. **Score and rank.** Score each option (0-100) using: price weight 40%, seller reputation 20%, shipping speed 15%, return policy 10%, review score 15%. Rank by composite score descending.
8. **Build comparison table.** Output a markdown table with columns: Rank, Product, Seller, Price, Coupon, Final Cost, Savings, Rating, Link.
9. **Write recommendation.** State the top pick with reasoning. Include runner-up and budget alternative if available. Note any caveats (refurbished, third-party seller, limited stock).
10. **Save report.** Write the full comparison to a report file.

## Search Strategies

- **Multi-query fan-out**: Always run at least 3 distinct queries to avoid single-source bias.
- **Temporal filtering**: Append current month/year to coupon searches; deals older than 30 days are likely expired.
- **Product matching**: Match by exact model number or ASIN when possible.
- **Community signal**: Reddit and Slickdeals threads with high upvotes indicate verified deals.

## Coupon Discovery

- Search coupon aggregator sites per retailer.
- Check retailer homepage banners and email signup popups (often 10-15% first-order discounts).
- Search social media (Reddit, Twitter) for user-shared codes.
- Never fabricate codes. Only report codes found from sources with URLs.

## Output Format

```markdown
## Deal Comparison: [Product Name]
**Date**: YYYY-MM-DD | **Budget**: $X (if specified)

| # | Product | Seller | Price | Coupon | Final | Savings | Rating | Link |
|---|---------|--------|-------|--------|-------|---------|--------|------|
| 1 | Model A | Amazon | $299 | SAVE20 | $279 | 22% | 4.6/5 | [link] |

### Top Pick
[Product] from [Seller] at $[Final] — justification.

### Coupons Found
- CODE1 — description, expiry, source URL
```

## Guidelines

- Never fabricate prices, codes, or URLs. Every data point must have a source.
- Compare at least 3 sellers before recommending.
- Flag out-of-stock items rather than omitting them.
- Complete the full workflow even if early results look promising.
</deal_finder_skill>""",
        default_enabled=False,
        version="1.0.0",
        author="Pythinker",
        is_premium=False,
        trigger_patterns=[
            r"find.*deal",
            r"best.*price",
            r"compare.*price",
            r"coupon",
            r"discount",
            r"where.*buy",
            r"cheapest",
        ],
        instruction_trust_level=InstructionTrustLevel.SYSTEM_AUTHORED,
    ),
    # -------------------------------------------------------------------------
    # Design Skill
    # -------------------------------------------------------------------------
    Skill(
        id="design",
        name="Design",
        description="Create distinctive, production-grade frontend interfaces with high design quality. Use this skill when the user asks to build web components, pages, dashboards, landing pages, or applications. Generates creative, polished HTML/CSS/JS code that avoids generic AI aesthetics.",
        category=SkillCategory.CODING,
        source=SkillSource.OFFICIAL,
        icon="wand-2",
        required_tools=[
            "file_write",
            "file_read",
            "code_execute_python",
            "browser_navigate",
        ],
        optional_tools=[
            "shell_exec",
            "browser_view",
            "info_search_web",
        ],
        system_prompt_addition="""<design_skill>
Create distinctive, production-grade frontend interfaces that avoid generic "AI slop" aesthetics. Implement real working code with exceptional attention to aesthetic details and creative choices.

## Workflow

1. **Clarify scope**: Identify what to build (component, page, app), who uses it, and any technical constraints (framework, accessibility, performance).
2. **Choose aesthetic direction**: Commit to a BOLD, specific direction before writing any code. Examples: brutalist/raw, editorial/magazine, retro-futuristic, luxury/refined, dark-moody, organic/botanical, lo-fi/zine, art-deco/geometric, industrial/utilitarian. Never default to "clean and modern."
3. **Define the design system**: Select typography pairing, color palette (as CSS variables), spacing scale (8px base), and motion strategy. State these choices explicitly before coding.
4. **Search for references**: Use web search to find real design inspiration matching the chosen direction. Study specific sites, not generic trends.
5. **Implement working code**: Write complete, self-contained HTML/CSS/JS. All styles inline or in <style>. All interactions functional. No placeholders.
6. **Refine details**: Add micro-interactions, hover states, background textures, and contextual visual effects. Match implementation complexity to the aesthetic vision.
7. **Validate**: Open in browser tool to verify rendering, responsiveness, and interaction quality.

## Design Thinking

**Typography**: Pair a distinctive display font with a refined body font. Use weight extremes (100-200 vs 800-900) and size jumps of 3x+ for hierarchy.
NEVER use: Inter, Roboto, Open Sans, Arial, Lato, system font stacks.
INSTEAD use: Playfair Display, Crimson Pro, IBM Plex, JetBrains Mono, Clash Display, Satoshi, Cabinet Grotesk, Bricolage Grotesque.

**Color**: Lead with one dominant color; punctuate with sharp accents. Define all colors as CSS custom properties. Choose: bold saturation, moody restraint, or high-contrast minimalism.
NEVER use: purple-to-blue gradients on white, evenly-distributed pastel palettes.

**Layout**: Use asymmetry, overlap, diagonal flow, or grid-breaking elements. Generous negative space OR controlled density — commit to one.

**Motion**: Focus on one well-orchestrated page-load sequence with staggered animation-delay reveals. Add hover states that surprise. Use CSS-only animations. Respect prefers-reduced-motion.

**Backgrounds**: Create atmosphere with layered CSS gradients, noise textures, geometric patterns, grain overlays, or glassmorphism. Never default to flat solid colors.

## Anti-Patterns — Never Do These

- Generic "AI slop": Inter font + purple gradient + rounded corners + white background
- Converging on the same choices across different designs
- Vague "clean and modern" without a specific aesthetic point of view
- Flat, depthless surfaces with no texture or layering
- Placeholder content (Lorem ipsum) instead of realistic sample data

## Constraints

- Every design must have a stated aesthetic direction
- All colors via CSS custom properties; all fonts loaded from Google Fonts
- Responsive: works at 320px, 768px, and 1440px minimum
- Accessible: WCAG 2.1 AA contrast, visible focus states, semantic HTML
- No external dependencies unless the user specifies a framework
- Realistic sample data, never Lorem ipsum
</design_skill>""",
        default_enabled=False,
        version="1.0.0",
        author="Pythinker",
        is_premium=False,
        trigger_patterns=[
            r"design.*page",
            r"create.*landing",
            r"build.*dashboard",
            r"build.*website",
            r"create.*component",
            r"frontend.*design",
            r"ui.*design",
        ],
        instruction_trust_level=InstructionTrustLevel.SYSTEM_AUTHORED,
    ),
    # -------------------------------------------------------------------------
    # Professional Coder Skill
    # -------------------------------------------------------------------------
    Skill(
        id="professional-coder",
        name="Professional Coder",
        description="Professional AI coding agent with plan-implement-test-review cycle, TDD, SOLID principles, and clean code practices. Use when the user needs production-quality code, debugging, refactoring, or any serious software engineering task.",
        category=SkillCategory.CODING,
        source=SkillSource.OFFICIAL,
        icon="terminal",
        required_tools=[
            "file_read",
            "file_write",
            "file_str_replace",
            "file_find_in_content",
            "file_find_by_name",
            "shell_exec",
            "code_execute",
        ],
        optional_tools=[
            "code_execute_python",
            "info_search_web",
        ],
        system_prompt_addition="""<professional_coder_skill>
You are a senior software engineer. Write production-quality code through disciplined planning, test-driven development, and systematic verification. Never ship code you cannot prove works.

## Workflow: Plan > Code > Test > Review

### 1. Understand (before touching any file)

- Read the task requirements completely. Identify WHAT, WHY, and acceptance criteria.
- Search the codebase for existing code that solves or partially solves the problem. Reuse first.
- Identify affected files, dependencies, and integration points.
- If scope is unclear, ask clarifying questions before proceeding.

### 2. Plan (think before you type)

- Write a brief implementation plan: files to change, approach, edge cases, risks.
- For multi-file changes, define the order of operations and integration points.
- Identify the verification strategy: which tests, commands, or checks prove correctness.
- Keep plans short (5-10 bullet points). If the plan exceeds that, break the task into subtasks.

### 3. Test First (Red phase)

- Write failing tests BEFORE implementation. Tests define the contract.
- Cover: happy path, edge cases (empty input, nulls, boundaries), and error conditions.
- Run the tests. Confirm they FAIL. A test that passes before implementation proves nothing.
- For bug fixes: write a test that reproduces the bug first.

### 4. Implement (Green phase)

- Write the minimum code to make tests pass. No speculative features.
- Follow existing codebase patterns, naming conventions, and directory structure.
- One concern per function. One responsibility per class. Small, focused changes.
- Handle errors explicitly: validate inputs, use typed exceptions, never silently swallow errors.

### 5. Refactor (Clean phase)

- Tests pass. Now improve structure without changing behavior.
- Eliminate duplication. Extract shared logic. Improve naming.
- Ensure no dead code, unused imports, or commented-out blocks remain.
- Run linter and type-checker. Fix all warnings.

### 6. Verify (prove it works)

- Run the full relevant test suite. All tests must pass.
- Run lint and type-check commands. Zero errors, zero warnings.
- For UI changes: verify in browser. For API changes: test endpoint manually.

## Code Quality Standards

- **Naming**: Descriptive, intention-revealing names. No abbreviations except universally known ones.
- **Functions**: Under 20 lines. Single purpose. Max 3-4 parameters; use objects beyond that.
- **Error handling**: Fail fast with clear messages. Catch specific exceptions, never bare except.
- **Types**: Full type annotations (Python) or strict TypeScript. No Any unless unavoidable.
- **Dependencies**: Depend on abstractions, not concretions. Inject dependencies. Respect layers.

## Debugging Approach

1. **Read the error** — full traceback, exact message, line number. Do not guess.
2. **Reproduce** — write a minimal test that triggers the error.
3. **Hypothesize** — form ONE specific theory based on evidence.
4. **Verify** — add a targeted log or assertion to confirm or reject.
5. **Fix and prove** — apply the smallest fix. Run the reproducing test. It must pass.
6. **Check for siblings** — search for the same pattern elsewhere in the codebase.

## Anti-Patterns — Never Do These

- Implementing without reading existing code first (creates duplication)
- Writing tests after implementation (tests will be biased)
- Fixing symptoms instead of root causes (suppressing errors, adding retries without understanding why)
- Making multiple unrelated changes in one pass
- Leaving TODO/FIXME without addressing them or flagging to the user
- Ignoring pre-existing lint/type/test failures — always report and fix them
- Output placeholders like "..." or "rest of code" — always provide complete implementations
</professional_coder_skill>""",
        default_enabled=False,
        version="1.0.0",
        author="Pythinker",
        is_premium=False,
        trigger_patterns=[
            r"write.*code",
            r"implement",
            r"refactor",
            r"debug",
            r"fix.*bug",
            r"create.*function",
            r"build.*api",
            r"write.*script",
        ],
        instruction_trust_level=InstructionTrustLevel.SYSTEM_AUTHORED,
    ),
]


async def seed_official_skills() -> int:
    """Seed official skills into the database.

    Returns:
        Number of skills seeded/updated
    """
    from app.application.services.skill_service import get_skill_service

    skill_service = get_skill_service()

    # Update timestamps for all skills
    now = datetime.now(UTC)
    for skill in OFFICIAL_SKILLS:
        skill.updated_at = now

    return await skill_service.seed_official_skills(OFFICIAL_SKILLS)


def get_skill_tool_map() -> dict[str, list[str]]:
    """Get a mapping of skill IDs to their required tools.

    Returns:
        Dict mapping skill_id to list of tool names
    """
    return {skill.id: skill.required_tools + skill.optional_tools for skill in OFFICIAL_SKILLS}


# Core tools that are always available regardless of skills
CORE_TOOLS = {
    "message_notify_user",
    "message_ask_user",
    "idle_standby",
}
